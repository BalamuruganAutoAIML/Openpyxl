"""
OpenPyXL Learning Examples
Practical demonstrations of common openpyxl operations
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
import csv
from datetime import datetime

# ============================================================================
# Example 1: Reading and Writing Basic Excel Files
# ============================================================================

def example_1_basic_read_write():
    """Create a simple Excel file and read it back"""
    print("\n" + "="*60)
    print("Example 1: Basic Read/Write Operations")
    print("="*60)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    
    # Write data
    ws['A1'] = "Name"
    ws['B1'] = "Age"
    ws['C1'] = "Email"
    
    ws.append(["John", 28, "john@example.com"])
    ws.append(["Jane", 32, "jane@example.com"])
    ws.append(["Bob", 25, "bob@example.com"])
    
    # Save
    wb.save("example_output_1.xlsx")
    print("[OK] Created: example_output_1.xlsx")
    
    # Read back
    wb = load_workbook("example_output_1.xlsx")
    ws = wb.active
    
    print("\nData from file:")
    for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=3, values_only=True):
        print(row)

# ============================================================================
# Example 2: Styling and Formatting
# ============================================================================

def example_2_styling():
    """Apply various formatting styles to cells"""
    print("\n" + "="*60)
    print("Example 2: Styling and Formatting")
    print("="*60)
    
    wb = Workbook()
    ws = wb.active
    
    # Style: Headers with color
    ws['A1'] = "Product"
    ws['B1'] = "Price"
    ws['C1'] = "Quantity"
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data
    data = [
        ["Laptop", 999.99, 5],
        ["Mouse", 29.99, 15],
        ["Keyboard", 79.99, 8],
    ]
    
    for row_data in data:
        ws.append(row_data)
    
    # Format price column as currency
    for row in range(2, 5):
        ws[f"B{row}"].number_format = '$#,##0.00'
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=4, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    
    wb.save("example_output_2.xlsx")
    print("[OK] Created: example_output_2.xlsx")
    print("  Features: Colors, Fonts, Currency Format, Borders, Column Widths")

# ============================================================================
# Example 3: Working with Multiple Sheets
# ============================================================================

def example_3_multiple_sheets():
    """Create workbook with multiple sheets and cross-references"""
    print("\n" + "="*60)
    print("Example 3: Multiple Sheets and Cross-References")
    print("="*60)
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Sheet 1: Sales Data
    ws1 = wb.create_sheet("Sales", 0)
    ws1.append(["Month", "Amount"])
    ws1.append(["January", 5000])
    ws1.append(["February", 6000])
    ws1.append(["March", 7000])
    
    # Sheet 2: Summary
    ws2 = wb.create_sheet("Summary", 1)
    ws2.append(["Metric", "Value"])
    ws2.append(["Total Sales", "=SUM(Sales!B2:B4)"])
    ws2.append(["Average", "=AVERAGE(Sales!B2:B4)"])
    ws2.append(["Max", "=MAX(Sales!B2:B4)"])
    
    # Format headers
    for cell in ws1[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Format as currency
    for row in range(2, 5):
        ws1[f"B{row}"].number_format = '$#,##0'
    
    for row in range(2, 5):
        ws2[f"B{row}"].number_format = '$#,##0'
    
    wb.save("example_output_3.xlsx")
    print("[OK] Created: example_output_3.xlsx")
    print("  Features: Multiple Sheets, Cross-Sheet Formulas, Automatic Calculations")

# ============================================================================
# Example 4: Using Formulas
# ============================================================================

def example_4_formulas():
    """Add various Excel formulas"""
    print("\n" + "="*60)
    print("Example 4: Using Formulas")
    print("="*60)
    
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = "Product"
    ws['B1'] = "Price"
    ws['C1'] = "Quantity"
    ws['D1'] = "Total"
    ws['E1'] = "Category"
    
    # Add data
    ws.append(["Laptop", 1000, 2, "=B2*C2", "Electronics"])
    ws.append(["Mouse", 25, 5, "=B3*C3", "Accessories"])
    ws.append(["Keyboard", 75, 3, "=B4*C4", "Accessories"])
    
    # Add summary formulas
    ws.append(["", "", "", "", ""])
    ws['A6'] = "Summary"
    ws['B6'] = "Total Items:"
    ws['C6'] = "=SUM(C2:C4)"
    ws['B7'] = "Total Value:"
    ws['C7'] = "=SUM(D2:D4)"
    ws['B8'] = "Average Price:"
    ws['C8'] = "=AVERAGE(B2:B4)"
    
    # IF formula
    ws['B9'] = "High Value Items:"
    ws['C9'] = "=COUNTIF(D2:D4,\">1000\")"
    
    # Format as currency
    for row in range(2, 5):
        ws[f"B{row}"].number_format = '$#,##0'
        ws[f"D{row}"].number_format = '$#,##0'
    
    ws['C7'].number_format = '$#,##0'
    ws['C8'].number_format = '$#,##0'
    
    wb.save("example_output_4.xlsx")
    print("[OK] Created: example_output_4.xlsx")
    print("  Formulas: SUM, AVERAGE, COUNTIF, Basic Arithmetic")

# ============================================================================
# Example 5: Data Validation
# ============================================================================

def example_5_data_validation():
    """Add data validation rules"""
    print("\n" + "="*60)
    print("Example 5: Data Validation")
    print("="*60)
    
    wb = Workbook()
    ws = wb.active
    
    # Form layout
    ws['A1'] = "Employee Entry Form"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:B1')
    
    # Labels
    ws['A3'] = "Name:"
    ws['A4'] = "Department:"
    ws['A5'] = "Performance Rating:"
    ws['A6'] = "Start Date:"
    
    # Input cells
    ws['B3'].border = Border(bottom=Side(style='thin'))
    ws['B4'].border = Border(bottom=Side(style='thin'))
    ws['B5'].border = Border(bottom=Side(style='thin'))
    ws['B6'].border = Border(bottom=Side(style='thin'))
    
    # Department dropdown
    dv_dept = DataValidation(
        type="list",
        formula1='"Sales,Marketing,IT,HR,Finance"',
        allow_blank=False
    )
    dv_dept.error = 'Please select a valid department'
    dv_dept.errorTitle = 'Invalid Department'
    ws.add_data_validation(dv_dept)
    dv_dept.add('B4')
    
    # Rating dropdown (1-5)
    dv_rating = DataValidation(
        type="whole",
        operator="between",
        formula1="1",
        formula2="5",
        allow_blank=False
    )
    dv_rating.error = 'Rating must be between 1 and 5'
    dv_rating.errorTitle = 'Invalid Rating'
    ws.add_data_validation(dv_rating)
    dv_rating.add('B5')
    
    # Date validation
    dv_date = DataValidation(
        type="date",
        operator="greaterThanOrEqual",
        formula1="2000-01-01"
    )
    ws.add_data_validation(dv_date)
    dv_date.add('B6')
    ws['B6'].number_format = 'YYYY-MM-DD'
    
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    
    wb.save("example_output_5.xlsx")
    print("[OK] Created: example_output_5.xlsx")
    print("  Features: List Dropdowns, Numeric Range, Date Validation, Error Messages")

# ============================================================================
# Example 6: Charts
# ============================================================================

def example_6_charts():
    """Create charts"""
    print("\n" + "="*60)
    print("Example 6: Creating Charts")
    print("="*60)
    
    wb = Workbook()
    ws = wb.active
    
    # Data
    ws['A1'] = "Month"
    ws['B1'] = "Sales"
    ws['C1'] = "Expenses"
    
    data = [
        ["January", 5000, 3000],
        ["February", 6000, 3200],
        ["March", 7000, 3500],
        ["April", 6500, 3300],
    ]
    
    for row_data in data:
        ws.append(row_data)
    
    # Format headers
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    # Format numbers
    for row in range(2, 6):
        ws[f"B{row}"].number_format = '$#,##0'
        ws[f"C{row}"].number_format = '$#,##0'
    
    # Create Bar Chart
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.title = "Sales vs Expenses"
    bar_chart.y_axis.title = "Amount ($)"
    bar_chart.x_axis.title = "Month"
    
    data_ref = Reference(ws, min_col=2, min_row=1, max_col=3, max_row=5)
    categories = Reference(ws, min_col=1, min_row=2, max_row=5)
    bar_chart.add_data(data_ref, titles_from_data=True)
    bar_chart.set_categories(categories)
    
    ws.add_chart(bar_chart, "E2")
    
    # Create Pie Chart (Q1 Sales)
    pie_chart = PieChart()
    pie_chart.title = "Q1 Sales Distribution"
    
    pie_data = Reference(ws, min_col=2, min_row=1, max_row=4)
    pie_chart.add_data(pie_data, titles_from_data=True)
    
    ws.add_chart(pie_chart, "E15")
    
    wb.save("example_output_6.xlsx")
    print("[OK] Created: example_output_6.xlsx")
    print("  Features: Bar Charts, Pie Charts, Chart Titles, Axis Labels")

# ============================================================================
# Example 7: Convert CSV to Excel
# ============================================================================

def example_7_csv_to_excel():
    """Read CSV file and convert to Excel"""
    print("\n" + "="*60)
    print("Example 7: Convert CSV to Excel with Formatting")
    print("="*60)
    
    # Create a sample CSV file
    csv_file = "sample_data.csv"
    with open(csv_file, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["Name", "Department", "Salary"])
        writer.writerow(["John Smith", "Sales", 50000])
        writer.writerow(["Jane Doe", "Marketing", 55000])
        writer.writerow(["Bob Johnson", "IT", 65000])
        writer.writerow(["Alice Brown", "HR", 48000])
    
    # Read CSV and write to Excel
    wb = Workbook()
    ws = wb.active
    
    with open(csv_file, 'r') as f:
        reader = csv.reader(f)
        for row in reader:
            ws.append(row)
    
    # Format headers
    header_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Format salary column
    for row in range(2, 6):
        ws[f"C{row}"].number_format = '$#,##0'
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    
    wb.save("example_output_7.xlsx")
    print("[OK] Created: example_output_7.xlsx")
    print("  Converted CSV to Excel with formatting")

# ============================================================================
# Example 8: Merge and Filter
# ============================================================================

def example_8_merge_cells():
    """Merge cells and create report layout"""
    print("\n" + "="*60)
    print("Example 8: Merging Cells and Report Layout")
    print("="*60)
    
    wb = Workbook()
    ws = wb.active
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = "QUARTERLY SALES REPORT"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 25
    
    # Subtitle
    ws.merge_cells('A2:C2')
    ws['A2'] = f"Report Date: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Data section
    ws['A4'] = "Product Summary"
    ws['A4'].font = Font(bold=True, size=12)
    
    ws.append(["Product", "Sales", "Commission"])
    for _ in range(3):
        ws.append(["", "", ""])
    
    ws['A5'] = "Product A"
    ws['B5'] = 50000
    ws['C5'] = "=B5*0.05"
    
    ws['A6'] = "Product B"
    ws['B6'] = 45000
    ws['C6'] = "=B6*0.05"
    
    ws['A7'] = "Product C"
    ws['B7'] = 55000
    ws['C7'] = "=B7*0.05"
    
    # Summary section
    ws['A9'] = "TOTALS"
    ws['A9'].font = Font(bold=True)
    ws['B9'] = "=SUM(B5:B7)"
    ws['C9'] = "=SUM(C5:C7)"
    
    # Format numbers
    for row in range(5, 10):
        ws[f"B{row}"].number_format = '$#,##0'
        ws[f"C{row}"].number_format = '$#,##0'
    
    wb.save("example_output_8.xlsx")
    print("[OK] Created: example_output_8.xlsx")
    print("  Features: Merged Cells, Professional Layout, Summary Section")

# ============================================================================
# Run All Examples
# ============================================================================

if __name__ == "__main__":
    print("\n" + "="*60)
    print("OpenPyXL Learning Examples")
    print("="*60)
    
    example_1_basic_read_write()
    example_2_styling()
    example_3_multiple_sheets()
    example_4_formulas()
    example_5_data_validation()
    example_6_charts()
    example_7_csv_to_excel()
    example_8_merge_cells()
    
    print("\n" + "="*60)
    print("All examples completed successfully!")
    print("Check example_output_*.xlsx files")
    print("="*60 + "\n")
